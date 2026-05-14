"""
_session_lib.py — helpers compartidos entre plan_session.py y
feedback_session.py.

Sin dependencias entre scripts CLI: todos los helpers viven acá. No es un
módulo público — el prefijo `_` señala que solo lo usan los otros scripts
del proyecto.
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
MASTER_PLAN = PROJECT_ROOT / "master_plan.md"
ADJUSTMENTS = PROJECT_ROOT / "plan_adjustments.md"
LEDGER = PROJECT_ROOT / "executed_volume.md"

# Manual data sources (Excel maintained by the athlete — no Garmin involved).
MANUAL_DIR = DATA_DIR / "manual"
BLOOD_XLSX = MANUAL_DIR / "blood.xlsx"
ANTHRO_XLSX = MANUAL_DIR / "anthropometry.xlsx"
BLOOD_RANGES_YML = MANUAL_DIR / "blood_reference_ranges.yml"
BLOOD_PANEL_MD = PROJECT_ROOT / "blood_panel.md"
BODY_COMP_MD = PROJECT_ROOT / "body_composition.md"

# LTHR-based zone thresholds (atleta-specific, ver CLAUDE.md / master_plan.md).
LTHR = 172
ZONE_BOUNDS = [
    ("Z1", 0, 120),
    ("Z2", 120, 135),
    ("Z3", 135, 155),
    ("Z4", 155, 172),
    ("Z5", 172, 999),
]

# Sentinel para marcar secciones de session.md que esperan feedback_session.py
PENDING_FEEDBACK = "[pendiente — completar con `feedback_session.py`]"


# ---------- date utils ----------

def day_dir(d: date) -> Path:
    return DATA_DIR / d.isoformat()


# ---------- garmin sync helper ----------

def run_sync(target: date, also_yesterday: bool = True) -> int:
    """Run garmin_sync.py for the target date (and optionally yesterday)."""
    start = (target - timedelta(days=1)) if also_yesterday else target
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "scripts" / "garmin_sync.py"),
        "--from", start.isoformat(),
        "--to", target.isoformat(),
    ]
    print(f"\n[sync] {' '.join(cmd)}\n")
    return subprocess.call(cmd, cwd=str(PROJECT_ROOT))


# ---------- wellness ----------

def load_wellness(target: date) -> dict | None:
    p = day_dir(target) / "wellness.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  ! could not read {p.name}: {e}")
        return None


def load_wellness_extended(target: date) -> dict | None:
    """Load `data/<date>/wellness_extended.json` (training_readiness,
    training_status, max_metrics, intensity_minutes, etc. — all the per-day
    metrics garmin_sync.py writes beyond the core wellness5).
    """
    p = day_dir(target) / "wellness_extended.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  ! could not read {p.name}: {e}")
        return None


def load_athlete_metrics(target: date) -> dict | None:
    """Load `data/<date>/athlete_metrics.json` (lactate_threshold,
    cycling_ftp, race_predictions, fitness_age, body_composition, devices,
    etc. — the longitudinal performance profile snapshot).
    """
    p = day_dir(target) / "athlete_metrics.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  ! could not read {p.name}: {e}")
        return None


def latest_athlete_metrics() -> tuple[date, dict] | None:
    """Find the most recent `data/<date>/athlete_metrics.json` and load it.

    Useful for the coach's bootstrap: even if today's snapshot doesn't
    exist yet, return whatever is most recent (typically yesterday or the
    last sync). Returns None if no snapshot exists at all.
    """
    if not DATA_DIR.exists():
        return None
    for child in sorted(DATA_DIR.iterdir(), reverse=True):
        if not child.is_dir():
            continue
        try:
            d = date.fromisoformat(child.name)
        except ValueError:
            continue
        p = child / "athlete_metrics.json"
        if p.exists():
            try:
                return d, json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
    return None


def metric_history(filename: str, field_path: str) -> list[tuple[str, object]]:
    """Iterate every `data/<date>/<filename>` and extract a value from a
    dot-separated path (e.g. 'race_predictions.time5K' or
    'fitness_age.fitnessAge').

    Returns a chronologically-sorted list of (date_iso, value) pairs,
    skipping days where the file or path is missing/null. Use this to
    plot long-term evolution of any metric across either wellness.json,
    wellness_extended.json or athlete_metrics.json.

    Examples:
        # From athlete_metrics.json
        metric_history('athlete_metrics.json', 'cycling_ftp.functionalThresholdPower')
        metric_history('athlete_metrics.json', 'race_predictions.time5K')
        metric_history('athlete_metrics.json', 'lactate_threshold.power.functionalThresholdPower')

        # From wellness_extended.json
        metric_history('wellness_extended.json', 'fitness_age.fitnessAge')
        metric_history('wellness_extended.json', 'training_status.mostRecentVO2Max.generic.vo2MaxPreciseValue')
    """
    keys = field_path.split(".")
    out: list[tuple[str, object]] = []
    if not DATA_DIR.exists():
        return out
    for child in sorted(DATA_DIR.iterdir()):
        if not child.is_dir():
            continue
        try:
            date.fromisoformat(child.name)
        except ValueError:
            continue
        p = child / filename
        if not p.exists():
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        cur: object = data
        for k in keys:
            if not isinstance(cur, dict) or k not in cur:
                cur = None
                break
            cur = cur[k]
        if cur is not None and cur != {} and cur != []:
            out.append((child.name, cur))
    return out


def wellness_summary_fields(target: date) -> dict:
    """Return the structured wellness summary used by both display and md write."""
    w = load_wellness(target) or {}
    sleep_dto = (w.get("sleep") or {}).get("dailySleepDTO") or {}
    score = (sleep_dto.get("sleepScores") or {}).get("overall", {}).get("value")
    dur_s = sleep_dto.get("sleepTimeSeconds")
    sleep_dur = "—"
    if dur_s:
        h, rem = divmod(int(dur_s), 3600)
        m, _ = divmod(rem, 60)
        sleep_dur = f"{h}h{m:02d}m"

    hrv_summary = (w.get("hrv") or {}).get("hrvSummary") or {}
    rhr_block = (
        ((w.get("resting_heart_rate") or {}).get("allMetrics") or {}).get("metricsMap") or {}
    ).get("WELLNESS_RESTING_HEART_RATE") or []
    rhr = rhr_block[0].get("value") if rhr_block else None

    bb_low = bb_high = None
    bb = w.get("body_battery") or []
    if isinstance(bb, list) and bb:
        try:
            arr = bb[0].get("bodyBatteryValuesArray") or []
            vals = [x[1] for x in arr if isinstance(x, list) and len(x) >= 2 and x[1] is not None]
            if vals:
                bb_low, bb_high = min(vals), max(vals)
        except Exception:
            pass

    stress = w.get("stress") or {}

    return {
        "sleep_score": score,
        "sleep_dur": sleep_dur,
        "hrv_avg": hrv_summary.get("lastNightAvg") if hrv_summary else None,
        "hrv_max": hrv_summary.get("lastNight5MinHigh") if hrv_summary else None,
        "hrv_status": hrv_summary.get("status") if hrv_summary else None,
        "rhr": rhr,
        "bb_low": bb_low,
        "bb_high": bb_high,
        "stress_avg": stress.get("avgStressLevel") if stress else None,
        "stress_max": stress.get("maxStressLevel") if stress else None,
    }


def print_wellness(target: date) -> None:
    w = load_wellness(target)
    print(f"\n=== Wellness {target.isoformat()} ===")
    if not w:
        print("  (sin wellness.json — corré `garmin_sync.py` para este día)")
        return
    f = wellness_summary_fields(target)
    rhr_str = f"{f['rhr']:.0f}" if f['rhr'] is not None else "—"
    print(f"  Sueño: score {f['sleep_score'] if f['sleep_score'] is not None else '—'},"
          f" duración {f['sleep_dur']}")
    print(f"  HRV (anoche): avg {f['hrv_avg'] if f['hrv_avg'] is not None else '—'} ms"
          f" / max {f['hrv_max'] if f['hrv_max'] is not None else '—'} ms"
          f" ({f['hrv_status'] or '—'})")
    print(f"  RHR: {rhr_str} bpm")
    print(f"  Body Battery (rango día): {f['bb_low'] if f['bb_low'] is not None else '—'} → "
          f"{f['bb_high'] if f['bb_high'] is not None else '—'}")
    print(f"  Estrés: avg {f['stress_avg'] if f['stress_avg'] is not None else '—'},"
          f" max {f['stress_max'] if f['stress_max'] is not None else '—'}")


def build_wellness_block(target: date) -> str:
    w = load_wellness(target)
    if not w:
        return "_Sin wellness.json para este día._"
    f = wellness_summary_fields(target)
    rhr_str = f"{f['rhr']:.0f}" if f['rhr'] is not None else "—"
    return (
        f"- Sueño: score {f['sleep_score'] if f['sleep_score'] is not None else '—'},"
        f" duración {f['sleep_dur']}\n"
        f"- HRV (anoche): avg {f['hrv_avg'] if f['hrv_avg'] is not None else '—'} ms"
        f" / max {f['hrv_max'] if f['hrv_max'] is not None else '—'} ms"
        f" ({f['hrv_status'] or '—'})\n"
        f"- RHR: {rhr_str} bpm\n"
        f"- Body Battery (rango día):"
        f" {f['bb_low'] if f['bb_low'] is not None else '—'} →"
        f" {f['bb_high'] if f['bb_high'] is not None else '—'}\n"
        f"- Estrés: avg {f['stress_avg'] if f['stress_avg'] is not None else '—'},"
        f" max {f['stress_max'] if f['stress_max'] is not None else '—'}"
    )


# ---------- master plan lookup ----------

def find_master_plan_target(target: date) -> str | None:
    """Best-effort: extract the row of master_plan.md whose row starts with
    | <Día> | YYYY-MM-DD | so the caller sees what was prescribed."""
    if not MASTER_PLAN.exists():
        return None
    text = MASTER_PLAN.read_text(encoding="utf-8")
    iso = target.isoformat()
    pat = re.compile(rf"^\|[^|]+\|\s*{iso}\s*\|.*$", re.MULTILINE)
    m = pat.search(text)
    return m.group(0).strip() if m else None


# ---------- activities + zones ----------

def list_today_activities(target: date) -> list[Path]:
    adir = day_dir(target) / "activities"
    if not adir.exists():
        return []
    return sorted(adir.glob("*.json"))


def hr_zone_of(hr: int | float | None) -> str | None:
    if hr is None:
        return None
    for name, lo, hi in ZONE_BOUNDS:
        if lo <= hr < hi:
            return name
    return None


def fmt_dur(seconds: float | int | None) -> str:
    if not seconds:
        return "—"
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}h{m:02d}m{sec:02d}s" if h else f"{m}m{sec:02d}s"


def parse_fit_zones(fit_path: Path) -> dict | None:
    """Distribute time across zones using consecutive timestamp deltas
    from the .fit's record messages."""
    try:
        from fitparse import FitFile
    except ImportError:
        return None
    try:
        ff = FitFile(str(fit_path))
        ff.parse()
    except Exception as e:
        print(f"  ! could not parse {fit_path.name}: {e}")
        return None

    seconds_in_zone = {z[0]: 0.0 for z in ZONE_BOUNDS}
    prev_ts = None
    prev_hr = None
    for msg in ff.get_messages("record"):
        d = {f.name: f.value for f in msg}
        ts = d.get("timestamp")
        hr = d.get("heart_rate")
        if prev_ts is not None and ts is not None and prev_hr is not None:
            dt = (ts - prev_ts).total_seconds()
            if 0 < dt < 30:
                z = hr_zone_of(prev_hr)
                if z:
                    seconds_in_zone[z] += dt
        prev_ts = ts
        prev_hr = hr
    return seconds_in_zone


def fit_files_block(target: date) -> str:
    paths = list_today_activities(target)
    if not paths:
        return "ninguno todavía"
    lines = []
    for p in paths:
        try:
            a = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        aid = a.get("activityId")
        name = a.get("activityName") or "(sin nombre)"
        sport = (a.get("activityType") or {}).get("typeKey") or ""
        dur = a.get("duration") or 0
        dist = a.get("distance") or 0
        avg_hr = a.get("averageHR")
        details = [f"{sport}", fmt_dur(dur)]
        if dist:
            details.append(f"{dist/1000:.2f} km")
        if avg_hr:
            details.append(f"FC media {int(avg_hr)} bpm")
        lines.append(
            f"- `data/{target.isoformat()}/activities/{aid}.fit` — {name} · "
            + " · ".join(details)
        )
    return "\n".join(lines)


def print_performance_feedback(target: date) -> None:
    """Print full quantitative feedback for the day's activities."""
    print(f"\n=== Performance feedback {target.isoformat()} ===")

    target_row = find_master_plan_target(target)
    if target_row:
        print("  Plan (master_plan.md):")
        print(f"    {target_row}")
    else:
        print("  Plan (master_plan.md): _no se encontró fila para esta fecha_")

    paths = list_today_activities(target)
    if not paths:
        print("  Sin actividades sincronizadas para este día.")
        return

    for p in paths:
        try:
            a = json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  ! could not read {p.name}: {e}")
            continue

        aid = a.get("activityId")
        name = a.get("activityName") or "(sin nombre)"
        sport = (a.get("activityType") or {}).get("typeKey") or ""
        dur = a.get("duration") or 0
        dist = a.get("distance") or 0
        avg_hr = a.get("averageHR")
        max_hr = a.get("maxHR")
        cad_avg = a.get("averageRunCadence") or a.get("averageCadence")

        print(f"\n  ─── activity {aid} · {sport} · {name!r} ───")
        print(f"    Duración:  {fmt_dur(dur)}")
        if dist:
            km = dist / 1000
            print(f"    Distancia: {km:.2f} km")
            if dur and km:
                pace_s = dur / km
                pm, ps = divmod(int(pace_s), 60)
                print(f"    Pace:      {pm}:{ps:02d} /km")
        if avg_hr:
            print(f"    FC media:  {avg_hr:.0f} bpm  ({hr_zone_of(avg_hr)})")
        if max_hr:
            print(f"    FC max:    {max_hr:.0f} bpm  ({hr_zone_of(max_hr)})")
        if cad_avg:
            print(f"    Cadencia:  {cad_avg:.0f}")

        fit_path = p.with_suffix(".fit")
        if fit_path.exists():
            zones = parse_fit_zones(fit_path)
            if zones and any(zones.values()):
                total = sum(zones.values()) or 1
                print("    Zonas (s · %):")
                for z, _, _ in ZONE_BOUNDS:
                    sec = zones[z]
                    pct = 100 * sec / total
                    if sec:
                        print(f"      {z}: {int(sec):4d}s ({pct:4.1f}%)")


# ---------- last sessions / alarms context ----------

ALARM_PATTERNS = re.compile(r"(?i)\b(alarma|dolor|molestia|rir|inflama|tendin|pinchaz)")


def last_n_sessions(n: int = 3, before: date | None = None) -> list[dict]:
    """Return up to N most recent activities (across all dirs) before `before`."""
    if not DATA_DIR.exists():
        return []
    pairs: list[tuple[str, Path]] = []
    for dd in sorted(DATA_DIR.iterdir()):
        if not dd.is_dir():
            continue
        try:
            d = date.fromisoformat(dd.name)
        except Exception:
            continue
        if before and d >= before:
            continue
        adir = dd / "activities"
        if not adir.exists():
            continue
        for path in sorted(adir.glob("*.json")):
            pairs.append((dd.name, path))
    out = []
    for day, path in pairs[-n:]:
        try:
            a = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        out.append({"day": day, "activity": a, "path": path})
    return out


def collect_alarms(target: date, days_back: int = 7) -> list[str]:
    alarms: list[str] = []
    if not DATA_DIR.exists():
        return alarms
    for dd in sorted(DATA_DIR.iterdir(), reverse=True):
        if not dd.is_dir():
            continue
        try:
            d = date.fromisoformat(dd.name)
        except Exception:
            continue
        if d > target or (target - d).days > days_back:
            continue
        for fn in ("notes.md", "session.md"):
            p = dd / fn
            if not p.exists():
                continue
            try:
                text = p.read_text(encoding="utf-8")
            except Exception:
                continue
            for line in text.splitlines():
                if ALARM_PATTERNS.search(line):
                    alarms.append(f"- [{dd.name} · {fn}] {line.strip()}")
    return alarms


def print_recent_context(target: date) -> None:
    """Show last 3 sessions + open alarms — context for plan_session."""
    print(f"\n=== Contexto reciente (al {target.isoformat()}) ===")
    sessions = last_n_sessions(3, before=target)
    if not sessions:
        print("  Últimas sesiones: _ninguna registrada antes de hoy._")
    else:
        print("  Últimas 3 sesiones:")
        for rec in sessions:
            a = rec["activity"]
            sport = (a.get("activityType") or {}).get("typeKey") or ""
            name = a.get("activityName") or "(sin nombre)"
            dur = a.get("duration") or 0
            dist = a.get("distance") or 0
            hr = a.get("averageHR")
            line = f"    {rec['day']} · {sport:25s} · {fmt_dur(dur):8s}"
            if dist:
                line += f" · {dist/1000:.2f} km"
            if hr:
                line += f" · FC {int(hr)}"
            line += f" · {name}"
            print(line)

    alarms = collect_alarms(target)
    if alarms:
        print("\n  Alarmas abiertas (últimos 7 días):")
        for line in alarms:
            print(f"    {line}")
    else:
        print("\n  Alarmas abiertas: ninguna detectada en últimos 7 días.")


# ---------- input loop ----------

def prompt_body_issues_loop(target: date) -> list[dict]:
    """Loop flexible de captura de molestias / cargas / lesiones.

    Muestra primero las issues actualmente abiertas según
    `executed_volume.md → Bitácora corporal` para que el atleta sepa qué
    está vivo. Después acepta N entradas (parte / severidad / estado /
    notas) hasta que se ingrese una parte vacía.

    Para cerrar una issue abierta, el atleta carga la misma parte con
    severidad 0 y estado=resolved (default cuando severidad=0).
    """
    print("\n=== Bitácora corporal ===")
    open_issues = current_open_body_issues()
    if open_issues:
        print("  Abiertas según el ledger:")
        for iss in sorted(open_issues, key=lambda x: x["parte"].lower()):
            note_short = (iss.get("notas") or "")[:60]
            print(f"    · {iss['parte']} (sev {iss['severidad']}, "
                  f"abierta desde {iss['fecha']}) — {note_short}")
        print("  Podés agregar nuevas, actualizar severidad de una existente, "
              "o cerrar una (severidad 0 + estado resolved).")
    else:
        print("  No hay molestias / lesiones abiertas en el ledger.")

    print("\nLogueá cada parte cargada/dolorida o resuelta hoy.")
    print("Enter vacío en la primera pregunta = terminar.")

    issues: list[dict] = []
    while True:
        try:
            parte = input("\n  Parte (ej. tibiales, hombro izq, lumbar) o Enter para terminar: ").strip()
        except EOFError:
            break
        if not parte:
            break

        sev = prompt_int_0_10(
            f"  Severidad de {parte}",
            "0 = ya no molesta / 5 = molestia moderada / 10 = dolor agudo o lesión activa.",
        )
        if sev == "":
            print("  ! sin severidad — entrada descartada.")
            continue

        sev_int = int(sev)
        default_estado = "resolved" if sev_int == 0 else "open"
        try:
            estado_raw = input(
                f"  Estado [open/resolved] (Enter = {default_estado}): "
            ).strip().lower()
        except EOFError:
            estado_raw = ""
        estado = estado_raw if estado_raw in ("open", "resolved") else default_estado

        try:
            notas = input("  Notas breves (1 línea, opcional): ").strip()
        except EOFError:
            notas = ""

        issues.append({
            "parte": parte,
            "severidad": sev,
            "estado": estado,
            "notas": notas,
        })
        print(f"  ✅ {parte} · sev {sev}/10 · {estado}")

    if issues:
        print(f"\n✅ Bitácora: {len(issues)} entrada(s) capturada(s).")
    else:
        print("\n✅ Bitácora: sin entradas hoy.")
    return issues


def prompt_int_0_10(label: str, prompt: str) -> str:
    """Pedir un entero 0-10. Vacío permitido (= '—'). Retry hasta input válido."""
    print(f"\n--- {label} ---")
    print(prompt)
    print("  (entero 0-10; Enter vacío para saltar)")
    while True:
        try:
            raw = input().strip()
        except EOFError:
            return ""
        if raw == "":
            return ""
        try:
            v = int(raw)
        except ValueError:
            print("  ! no es un entero válido. Intentá de nuevo (o Enter para saltar).")
            continue
        if not 0 <= v <= 10:
            print("  ! fuera de rango 0-10. Intentá de nuevo.")
            continue
        return str(v)


def prompt_section(label: str, prompt: str, multiline_hint: str = "") -> str:
    """Read user input. Doble Enter para enviar (multilínea soportado)."""
    print(f"\n--- {label} ---")
    print(prompt)
    if multiline_hint:
        print(f"  ({multiline_hint})")
    print("  (terminar con una línea en blanco; doble Enter para enviar)")
    lines: list[str] = []
    while True:
        try:
            line = input()
        except EOFError:
            break
        if line == "":
            if lines:
                break
            continue
        lines.append(line)
    return "\n".join(lines).strip()


# ---------- session.md read/write ----------

SECTION_HEADERS = {
    "plan_original": "## Plan original",
    "plan_modificado": "## Plan modificado (pre-sesión)",
    "razon_ajuste": "## Razón del ajuste pre-sesión",
    "ejecutado": "## Sesión ejecutada",
    "desviaciones": "## Desviaciones durante la ejecución",
    "comentarios": "## Comentarios del atleta",
    "marcadores": "## Marcadores post-sesión",
    "wellness": "## Wellness pre-sesión",
    "fits": "## Archivos .fit asociados",
    "vinculo": "## Vínculo con master_plan.md y plan_adjustments.md",
}

SECTION_DOC = {
    "plan_original": "> Sesión tal como estaba en master_plan.md ANTES de cualquier modificación pre-sesión.",
    "plan_modificado": "> Sesión ajustada en función del wellness del día y/o eventos del calendario, ANTES de ejecutar.",
    "ejecutado": '> Lo que el atleta efectivamente realizó. Puede coincidir con "Plan modificado" o haber sufrido más cambios sobre la marcha.',
    "comentarios": "> Sensaciones, dolor, fatiga, contexto de vida que afecte la lectura de la sesión.",
    "marcadores": "> RPE = esfuerzo percibido global de la sesión (0-10). Bitácora = una entrada por cada parte del cuerpo cargada/dolorida/lesionada o resuelta hoy. La lista completa con histórico vive en `executed_volume.md` → sección `Bitácora corporal`.",
}


def session_md_path(target: date) -> Path:
    return day_dir(target) / "session.md"


def render_session_md(target: date, sections: dict) -> str:
    """Render the canonical session.md from a sections dict.

    Missing sections render as either PENDING_FEEDBACK (for Q4-Q6 fields if
    plan-only) or a clear placeholder.
    """
    target_row = find_master_plan_target(target) or "_no encontrado en master_plan.md_"
    iso = target.isoformat()

    def block(key: str, body: str | None) -> str:
        h = SECTION_HEADERS[key]
        doc = SECTION_DOC.get(key)
        body = body if body else "[awaiting user input]"
        return f"{h}\n\n" + (f"{doc}\n\n" if doc else "") + body + "\n"

    rpe = sections.get("rpe", "")
    body_issues = sections.get("body_issues") or []
    markers_lines = [f"- RPE (esfuerzo percibido global): **{rpe or '—'}** /10"]
    if body_issues:
        markers_lines.append("")
        markers_lines.append("**Bitácora corporal de hoy:**")
        for iss in body_issues:
            parte = iss.get("parte", "")
            sev = iss.get("severidad", "—")
            estado = iss.get("estado", "")
            notas = (iss.get("notas") or "").strip()
            line = f"- {parte} · sev {sev}/10 · {estado}"
            if notas:
                line += f" · {notas}"
            markers_lines.append(line)
    else:
        markers_lines.append("")
        markers_lines.append("_Sin entradas en la bitácora corporal hoy._")
    markers_body = "\n".join(markers_lines)

    parts = [
        f"# Sesión {iso}\n",
        block("plan_original", sections.get("plan_original")),
        block("plan_modificado", sections.get("plan_modificado")),
        f"{SECTION_HEADERS['razon_ajuste']}\n\n"
        + (sections.get("razon_ajuste") or "N/A") + "\n",
        block("ejecutado", sections.get("ejecutado")),
        f"{SECTION_HEADERS['desviaciones']}\n\n"
        + (sections.get("desviaciones") or "ninguna") + "\n",
        block("comentarios", sections.get("comentarios")),
        f"{SECTION_HEADERS['marcadores']}\n\n"
        f"{SECTION_DOC['marcadores']}\n\n"
        f"{markers_body}\n",
        f"{SECTION_HEADERS['wellness']}\n\n{build_wellness_block(target)}\n\n"
        f"_Auto-populated from `data/{iso}/wellness.json`._\n",
        f"{SECTION_HEADERS['fits']}\n\n{fit_files_block(target)}\n",
        f"{SECTION_HEADERS['vinculo']}\n\n"
        f"- **Master plan reference:** `{target_row}`\n"
        f"- **Adjustment log entry:** "
        f"Ver `plan_adjustments.md` — entry con `Date: {iso}`.\n",
    ]
    return "\n".join(parts)


SECTION_RE = re.compile(r"^##\s+(.+)$", re.MULTILINE)
RPE_RE = re.compile(r"RPE[^*\d]*\*\*\s*(\d{1,2})\s*\*\*", re.IGNORECASE)

# Bullets de bitácora dentro del bloque de marcadores en session.md, formato:
#   - {parte} · sev {N}/10 · {estado} · {notas}
ISSUE_BULLET_RE = re.compile(
    r"^\s*-\s*(?P<parte>[^·\n]+?)\s*·\s*sev\s*(?P<sev>\d{1,2})\s*/\s*10\s*·\s*"
    r"(?P<estado>open|resolved)(?:\s*·\s*(?P<notas>.*))?$",
    re.MULTILINE | re.IGNORECASE,
)

# Filas de la tabla 'Bitácora corporal' en executed_volume.md, formato:
#   | YYYY-MM-DD | parte | N | open|resolved | notas |
LEDGER_BITACORA_HEADER = "## Bitácora corporal"
LEDGER_BITACORA_ROW_RE = re.compile(
    r"^\|\s*(?P<fecha>\d{4}-\d{2}-\d{2})\s*\|\s*(?P<parte>[^|]+?)\s*\|\s*"
    r"(?P<sev>\d{1,2})\s*\|\s*(?P<estado>open|resolved)\s*\|\s*(?P<notas>[^|]*?)\s*\|\s*$",
    re.MULTILINE | re.IGNORECASE,
)
LEDGER_RPE_HEADER = "## RPE por día"


def parse_session_md(target: date) -> dict:
    """Read existing session.md and return a sections dict (best-effort).

    Extra: cuando encuentra la sección 'Marcadores post-sesión', parsea
    rpe (string '0'..'10') y body_issues (lista de dicts con parte /
    severidad / estado / notas).
    """
    p = session_md_path(target)
    if not p.exists():
        return {}
    text = p.read_text(encoding="utf-8")

    sections: dict = {}
    matches = list(SECTION_RE.finditer(text))
    for i, m in enumerate(matches):
        title = m.group(1).strip()
        for key, header in SECTION_HEADERS.items():
            if header[3:].strip() == title:
                start = m.end()
                end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
                body = text[start:end].strip()
                lines = body.splitlines()
                while lines and lines[0].startswith(">"):
                    lines.pop(0)
                while lines and lines[0].strip() == "":
                    lines.pop(0)
                body_clean = "\n".join(lines).strip()
                if body_clean in (PENDING_FEEDBACK, "[awaiting user input]", "N/A", "ninguna"):
                    continue
                sections[key] = body_clean
                if key == "marcadores":
                    if (mm := RPE_RE.search(body_clean)):
                        sections["rpe"] = mm.group(1)
                    issues = []
                    for im in ISSUE_BULLET_RE.finditer(body_clean):
                        issues.append({
                            "parte": im.group("parte").strip(),
                            "severidad": im.group("sev"),
                            "estado": im.group("estado").lower(),
                            "notas": (im.group("notas") or "").strip(),
                        })
                    if issues:
                        sections["body_issues"] = issues
                break
    return sections


def write_session_md(target: date, sections: dict) -> Path:
    """Write session.md with the given sections, creating dir if needed."""
    out = session_md_path(target)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(render_session_md(target, sections), encoding="utf-8")
    return out


# ---------- plan_adjustments.md / executed_volume.md updates ----------

def append_plan_adjustment(target: date, sections: dict) -> None:
    iso = target.isoformat()
    fit_paths = list_today_activities(target)
    fits = " + ".join(
        f"data/{iso}/activities/{p.stem}.fit" for p in fit_paths
    ) or "(sin .fit)"

    plan_orig = (sections.get("plan_original") or "").splitlines()
    plan_orig_first = plan_orig[0] if plan_orig else "(ver session.md)"

    entry = f"""

---
Date: {iso}
Original session: {plan_orig_first}
Modified to: {sections.get('plan_modificado','') or 'sin modificación'}
Executed as: {sections.get('ejecutado','') or '(ver session.md)'}
Reason: {sections.get('razon_ajuste','') or 'N/A — sin modificación pre-sesión'}
Source: data/{iso}/session.md + data/{iso}/wellness.json + {fits}
---
"""
    with open(ADJUSTMENTS, "a", encoding="utf-8") as f:
        f.write(entry)


def _append_to_section(header: str, ensure_block: str | None, rows_text: str) -> None:
    """Insert `rows_text` at the END of the section that starts with `header`.

    The "end of the section" is just before the next `^## ` header (or EOF).
    This is the fix for the bug where every helper appended at the end of
    the file, causing RPE and activity rows to accidentally land inside
    the Bitácora corporal section (the last section in the file).

    Behavior:
    - If the section header is missing AND `ensure_block` is provided, the
      block is appended to the end of the file plus the new rows. Use this
      when first-creating a section.
    - If the section header is missing AND `ensure_block` is None, this
      function is a no-op (caller asked to append into a section that
      doesn't exist, no action taken).
    """
    if not LEDGER.exists():
        return
    text = LEDGER.read_text(encoding="utf-8")
    if header not in text:
        if ensure_block is None:
            return
        with open(LEDGER, "a", encoding="utf-8") as f:
            f.write(ensure_block)
            f.write(rows_text)
            if not rows_text.endswith("\n"):
                f.write("\n")
        return

    header_pos = text.find(header)
    after_header = header_pos + len(header)
    next_header = re.search(r"\n## ", text[after_header:])
    section_end = after_header + next_header.start() if next_header else len(text)

    before = text[:section_end].rstrip("\n")
    after_text = text[section_end:]
    rows_clean = rows_text.rstrip("\n")

    new_text = before + "\n" + rows_clean + "\n"
    if after_text:
        new_text += "\n" + after_text.lstrip("\n")

    LEDGER.write_text(new_text, encoding="utf-8")


def append_ledger_rows(target: date) -> None:
    """Append a row to executed_volume.md for each activity of the day.

    Each row goes into the ISO-week section of `target` (e.g. `## 2026-W18`).
    If the week section doesn't exist yet, it gets created at the end of
    the file with the canonical header + table header.
    """
    paths = list_today_activities(target)
    if not paths:
        return
    if not LEDGER.exists():
        return
    iso = target.isoformat()
    iso_year, iso_week, _ = target.isocalendar()
    week_header = f"## {iso_year}-W{iso_week:02d}"

    monday = target - timedelta(days=target.isoweekday() - 1)
    sunday = monday + timedelta(days=6)

    lines: list[str] = []
    lines.append(f"<!-- feedback_session.py: append for {iso} at "
                 f"{datetime.utcnow().isoformat()}Z -->")
    for p in paths:
        try:
            a = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        sport = (a.get("activityType") or {}).get("typeKey") or ""
        dur = a.get("duration") or 0
        dist = a.get("distance") or 0
        avg_hr = a.get("averageHR")
        max_hr = a.get("maxHR")
        h, rem = divmod(int(dur), 3600)
        m, _ = divmod(rem, 60)
        dur_s = f"{h}h{m:02d}m" if h else f"{m}m"
        dist_s = f"{dist/1000:.2f} km" if dist else "—"
        hr_s = f"{int(avg_hr)} / {int(max_hr) if max_hr else '—'}" if avg_hr else "— / —"
        lines.append(
            f"| {iso} | {sport} | {dur_s} | {dist_s} | {hr_s} | _agregado por feedback_session_ |"
        )

    rows_text = "\n".join(lines)
    ensure_block = (
        f"\n\n{week_header} ({monday.strftime('%d/%m')} → {sunday.strftime('%d/%m')})\n\n"
        f"| Fecha | Modalidad | Duración | Distancia | FC media / max | Notas |\n"
        f"|---|---|---:|---:|---:|---|\n"
    )
    _append_to_section(week_header, ensure_block, rows_text)


def append_rpe_row(target: date, sections: dict) -> None:
    """Append a row to the 'RPE por día' table in executed_volume.md.

    Always lands inside the RPE section, regardless of which sections come
    after it in the file.
    """
    if not LEDGER.exists():
        return
    iso = target.isoformat()
    rpe = (sections.get("rpe") or "—").strip() or "—"
    note = (sections.get("comentarios") or "").splitlines()
    note_one = (note[0] if note else "").strip()[:80]

    ensure_block = (
        "\n\n---\n\n" + LEDGER_RPE_HEADER + "\n\n"
        "> RPE = esfuerzo percibido global de la sesión, escala 1-10\n"
        "> (1 = paseo, 5 = moderado, 8 = duro, 10 = al límite). Cargado por\n"
        "> `feedback_session.py` post-sesión.\n\n"
        "| Fecha | RPE | Notas |\n"
        "|---|---:|---|\n"
    )
    rows_text = f"| {iso} | {rpe} | {note_one} |"
    _append_to_section(LEDGER_RPE_HEADER, ensure_block, rows_text)


def append_body_issue_rows(target: date, issues: list[dict]) -> None:
    """Append one row per body issue to 'Bitácora corporal' in
    executed_volume.md. Always lands inside the Bitácora section.
    """
    if not LEDGER.exists() or not issues:
        return
    iso = target.isoformat()

    ensure_block = (
        "\n\n---\n\n" + LEDGER_BITACORA_HEADER + "\n\n"
        "> Append-only. Cada observación de carga / molestia / lesión va en una\n"
        "> fila nueva. Para cerrar una molestia, agregá fila con `estado=resolved`.\n"
        "> Para reportar empeoramiento, fila nueva con severidad mayor. La parte\n"
        "> es texto libre — escribí lo que tenga sentido (e.g. `tibial der`,\n"
        "> `hombro izq`, `cuádriceps`, `lumbar`).\n\n"
        "| Fecha | Parte | Severidad | Estado | Notas |\n"
        "|---|---|---:|---|---|\n"
    )
    rows = []
    for iss in issues:
        parte = (iss.get("parte") or "").strip().replace("|", "/")
        sev = (str(iss.get("severidad") or "")).strip() or "—"
        estado = (iss.get("estado") or "open").strip().lower()
        notas = (iss.get("notas") or "").strip().replace("|", "/")[:120]
        rows.append(f"| {iso} | {parte} | {sev} | {estado} | {notas} |")
    rows_text = "\n".join(rows)
    _append_to_section(LEDGER_BITACORA_HEADER, ensure_block, rows_text)


def read_bitacora_rows() -> list[dict]:
    """Devuelve TODAS las filas de la 'Bitácora corporal' como lista de dicts.

    Sirve para que el coach (o el script) sepa cuáles partes están
    actualmente abiertas (=última fila por parte con estado=open) y cuáles
    se resolvieron y cuándo.
    """
    if not LEDGER.exists():
        return []
    text = LEDGER.read_text(encoding="utf-8")
    if LEDGER_BITACORA_HEADER not in text:
        return []
    # buscar solo filas que vengan después del header
    start = text.index(LEDGER_BITACORA_HEADER)
    sub = text[start:]
    rows = []
    for m in LEDGER_BITACORA_ROW_RE.finditer(sub):
        rows.append({
            "fecha": m.group("fecha"),
            "parte": m.group("parte").strip(),
            "severidad": m.group("sev"),
            "estado": m.group("estado").lower(),
            "notas": m.group("notas").strip(),
        })
    return rows


def current_open_body_issues() -> list[dict]:
    """Para cada parte presente en la bitácora, devuelve la última observación
    cronológica solo si su estado es `open`. Es la 'foto' del estado actual
    del cuerpo del atleta."""
    rows = read_bitacora_rows()
    if not rows:
        return []
    # ordenar cronológicamente; si hay empate de fecha, la última en el
    # archivo gana (suficiente como heurística — el script no agrega
    # múltiples filas para la misma parte el mismo día sin querer).
    by_part: dict[str, dict] = {}
    for r in rows:
        key = r["parte"].lower()
        prev = by_part.get(key)
        if prev is None or r["fecha"] >= prev["fecha"]:
            by_part[key] = r
    return [r for r in by_part.values() if r["estado"] == "open"]


# ---------- manual data: blood + anthropometry ----------
#
# These two workbooks live at `data/manual/{blood,anthropometry}.xlsx` and
# are maintained by hand (the atleta drops in lab results / antropometrist
# evals). Format is wide — rows are markers/variables, columns are dates —
# because that's how the atleta receives the data from lab and evaluator.
# The readers below pivot to long format on the way out so callers get a
# normalized `list[dict]` (same shape style as `read_bitacora_rows()`).
#
# **Marker/variable names are preserved verbatim** (typos, parenthetical
# units, double-spaces all left intact). Canonical-name resolution is an
# interpretation-layer concern, not a reader concern.

def _coerce_excel_date(v: object) -> date | None:
    """Coerce a workbook cell to a `date`. Accepts:

    - real `datetime` / `date` objects (Excel-native date cells)
    - strings in `DD/MM/YYYY` or `D/M/YYYY` (e.g. '21/03/2019', '28/6/2021')
    - strings in `YYYY-MM-DD` or `DD-MM-YYYY`

    Returns None for anything else (blanks, numbers, unparseable strings).
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _coerce_excel_float(v: object) -> float | None:
    """Coerce a workbook cell to a float, or None if blank/unparseable.
    Accepts ints, floats, and numeric strings with `,` or `.` decimal."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _detect_date_header_row(ws, max_scan: int = 5) -> tuple[int, list[date | None]]:
    """Find the first row in the sheet whose cells from col B onwards are
    parseable as dates. Returns (row_number, [date|None per column from B]).

    Tolerant to a leading empty row (the blood workbook has r1 empty + r2
    dates; the anthropometry workbook has dates already at r1).

    Raises:
        ValueError if no qualifying row is found within `max_scan` rows.
    """
    upper = min(max_scan, ws.max_row)
    for r in range(1, upper + 1):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        if len(row) < 2:
            continue
        dates = [_coerce_excel_date(v) for v in row[1:]]
        if sum(1 for d in dates if d is not None) >= 2:
            return r, dates
    raise ValueError(
        f"No se encontró fila de encabezado con fechas en las primeras "
        f"{max_scan} filas de la hoja {ws.title!r}. Esperado: una fila "
        f"con fechas en columnas B en adelante (datetime nativo o string "
        f"DD/MM/YYYY)."
    )


def load_blood_panel(path: Path = BLOOD_XLSX) -> list[dict]:
    """Read the blood-panel workbook and pivot to long format.

    Returns:
        Sorted list of rows shaped:
            [{'fecha': 'YYYY-MM-DD',
              'marker': '<verbatim>',
              'valor': float,
              'sheet': '<sheet name>'}, ...]

    Schema expected:
        Sheet1 — wide: rows are markers (col A from header_row+1 onwards),
            columns are dates (cols B onwards in the auto-detected header
            row). Values are floats; empty cells skipped.
        Sheet2 — optional, transposed: row 2 has markers in cols B onwards,
            col A from row 3 has dates, cells are values. Merged into
            Sheet1 with Sheet1 winning on conflict `(fecha, marker)`.

    Markers preserved verbatim — typos and unit suffixes left intact.

    Raises:
        FileNotFoundError if `path` doesn't exist.
        ValueError with a Spanish hint if the sheet structure can't be parsed.
        RuntimeError if openpyxl is not installed.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"No existe {path}. Esperado: workbook con análisis de sangre "
            f"en formato ancho (fechas en columnas, marcadores en filas)."
        )
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl no está instalado. Corré: "
            "`.venv/bin/pip install openpyxl==3.1.5`"
        ) from e

    wb = load_workbook(path, data_only=True)
    if not wb.sheetnames:
        raise ValueError(f"{path}: el workbook no tiene hojas.")

    rows: list[dict] = []

    # ---- Sheet1: wide, rows-as-markers × cols-as-dates ----
    sheet1 = wb[wb.sheetnames[0]]
    header_row, dates = _detect_date_header_row(sheet1)
    for r in range(header_row + 1, sheet1.max_row + 1):
        marker_cell = sheet1.cell(row=r, column=1).value
        if marker_cell is None:
            continue
        marker = str(marker_cell).strip()
        if not marker:
            continue
        for col_idx, d in enumerate(dates, start=2):
            if d is None:
                continue
            v = _coerce_excel_float(sheet1.cell(row=r, column=col_idx).value)
            if v is None:
                continue
            rows.append({
                "fecha": d.isoformat(),
                "marker": marker,
                "valor": v,
                "sheet": sheet1.title,
            })

    # ---- Sheet2 (optional): transposed, dates-as-rows × markers-as-cols ----
    if len(wb.sheetnames) > 1:
        sheet2 = wb[wb.sheetnames[1]]
        if sheet2.max_row > 1 and sheet2.max_column > 1:
            marker_row = None
            for r in range(1, min(4, sheet2.max_row) + 1):
                row_vals = [
                    sheet2.cell(row=r, column=c).value
                    for c in range(2, sheet2.max_column + 1)
                ]
                if sum(1 for v in row_vals if isinstance(v, str) and v.strip()) >= 2:
                    marker_row = r
                    break
            if marker_row is not None:
                markers = []
                for c in range(2, sheet2.max_column + 1):
                    v = sheet2.cell(row=marker_row, column=c).value
                    markers.append(
                        str(v).strip() if isinstance(v, str) and v.strip() else None
                    )
                seen = {(r["fecha"], r["marker"]) for r in rows}
                for r in range(marker_row + 1, sheet2.max_row + 1):
                    d = _coerce_excel_date(sheet2.cell(row=r, column=1).value)
                    if d is None:
                        continue
                    for col_idx, marker in enumerate(markers, start=2):
                        if marker is None:
                            continue
                        v = _coerce_excel_float(
                            sheet2.cell(row=r, column=col_idx).value
                        )
                        if v is None:
                            continue
                        key = (d.isoformat(), marker)
                        if key in seen:
                            continue  # Sheet1 wins (per Phase 1 decision)
                        rows.append({
                            "fecha": d.isoformat(),
                            "marker": marker,
                            "valor": v,
                            "sheet": sheet2.title,
                        })
                        seen.add(key)

    rows.sort(key=lambda r: (r["fecha"], r["marker"]))
    return rows


def load_anthropometry(path: Path = ANTHRO_XLSX) -> list[dict]:
    """Read the anthropometry workbook and pivot to long format.

    Returns:
        Sorted list of rows shaped:
            [{'fecha': 'YYYY-MM-DD',
              'variable': '<verbatim>',
              'valor': float}, ...]

    Schema expected:
        Sheet1 only — wide: rows are variables (col A from header_row+1
            onwards), columns are dates (cols B onwards in the auto-detected
            header row). Sheet2+ ignored.

    Variables preserved verbatim — typos like 'Pr. Brazo Relaiado (cm)' and
    'Masa Osea (%]' are NOT cleaned here.

    Raises:
        FileNotFoundError if `path` doesn't exist.
        ValueError if Sheet1 structure can't be parsed.
        RuntimeError if openpyxl is not installed.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"No existe {path}. Esperado: workbook con antropometrías en "
            f"formato ancho (fechas en columnas, variables en filas)."
        )
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl no está instalado. Corré: "
            "`.venv/bin/pip install openpyxl==3.1.5`"
        ) from e

    wb = load_workbook(path, data_only=True)
    if not wb.sheetnames:
        raise ValueError(f"{path}: el workbook no tiene hojas.")

    sheet1 = wb[wb.sheetnames[0]]
    header_row, dates = _detect_date_header_row(sheet1)

    rows: list[dict] = []
    for r in range(header_row + 1, sheet1.max_row + 1):
        var_cell = sheet1.cell(row=r, column=1).value
        if var_cell is None:
            continue
        variable = str(var_cell).strip()
        if not variable:
            continue
        for col_idx, d in enumerate(dates, start=2):
            if d is None:
                continue
            v = _coerce_excel_float(sheet1.cell(row=r, column=col_idx).value)
            if v is None:
                continue
            rows.append({
                "fecha": d.isoformat(),
                "variable": variable,
                "valor": v,
            })

    rows.sort(key=lambda r: (r["fecha"], r["variable"]))
    return rows


def latest_blood_panel() -> tuple[date, list[dict]] | None:
    """Return (most-recent-draw date, rows of that draw only).

    Mirrors `latest_athlete_metrics()`: returns None if the workbook is
    missing or has no parseable data.
    """
    if not BLOOD_XLSX.exists():
        return None
    try:
        all_rows = load_blood_panel(BLOOD_XLSX)
    except (ValueError, FileNotFoundError):
        return None
    if not all_rows:
        return None
    latest = max(r["fecha"] for r in all_rows)
    return date.fromisoformat(latest), [r for r in all_rows if r["fecha"] == latest]


def latest_anthropometry() -> tuple[date, list[dict]] | None:
    """Return (most-recent-eval date, rows of that eval only).
    Mirrors `latest_athlete_metrics()`."""
    if not ANTHRO_XLSX.exists():
        return None
    try:
        all_rows = load_anthropometry(ANTHRO_XLSX)
    except (ValueError, FileNotFoundError):
        return None
    if not all_rows:
        return None
    latest = max(r["fecha"] for r in all_rows)
    return date.fromisoformat(latest), [r for r in all_rows if r["fecha"] == latest]


def blood_marker_history(marker: str) -> list[tuple[str, float]]:
    """All values for one blood marker across every draw, chronologically.

    `marker` must match the verbatim string used in the workbook (e.g.
    'Ferritina', 'Vitamina D', 'Hemoglobina glicosilada (%)'). No fuzzy
    matching — use exactly what's in the file. Returns
    `[('YYYY-MM-DD', value), ...]` ascending. Empty list if the marker is
    unknown or the workbook is missing.

    Mirrors `metric_history()` for Garmin JSON metrics.
    """
    if not BLOOD_XLSX.exists():
        return []
    try:
        all_rows = load_blood_panel(BLOOD_XLSX)
    except (ValueError, FileNotFoundError):
        return []
    return sorted(
        ((r["fecha"], r["valor"]) for r in all_rows if r["marker"] == marker),
        key=lambda t: t[0],
    )


def anthropometry_variable_history(variable: str) -> list[tuple[str, float]]:
    """All values for one anthropometry variable across every eval.

    `variable` must match verbatim (e.g. 'Peso (kg)', 'Masa Adiposa (%)',
    'Pl. Tricipital (mm)'). Returns `[('YYYY-MM-DD', value), ...]`.
    Mirrors `metric_history()` for Garmin JSON metrics.
    """
    if not ANTHRO_XLSX.exists():
        return []
    try:
        all_rows = load_anthropometry(ANTHRO_XLSX)
    except (ValueError, FileNotFoundError):
        return []
    return sorted(
        ((r["fecha"], r["valor"]) for r in all_rows if r["variable"] == variable),
        key=lambda t: t[0],
    )


# ---------- manual data: interpretation + persistence ----------
#
# Layered on top of the readers above. The flow is:
#
#   blood.xlsx + blood_reference_ranges.yml + active profile
#       → interpret_blood_panel()    → dict (categorías, flags, tendencias)
#       → refresh_blood_panel_md()   → blood_panel.md (living doc en root)
#
#   anthropometry.xlsx + heurísticas inline + active profile
#       → interpret_anthropometry()  → dict (bloques, flags, tendencias)
#       → refresh_body_composition_md() → body_composition.md
#
# Reglas:
# - Banderas DURAS son universales — vienen del `trigger_si` del YAML
#   (blood) o de chequeos hard-coded (anthro). Son señales clínicas, no
#   de performance.
# - Banderas BLANDAS son profile-aware — vienen de estar fuera de la
#   `target_atleta[<perfil>]` band del YAML (blood) o de targets por
#   perfil hard-coded (anthro).
# - Las "lecturas de entrenamiento" son seed escritas en español; editalas
#   directo en el YAML cuando no encajen con tu caso.

def _active_coach_profile() -> str:
    """Resolve the active coach profile name from profile.yml.

    Inlined here (instead of importing from `profiles.registry`) so that
    `_session_lib.py` stays self-contained and works from `python -c`
    invocations that haven't put `profiles/` on the path.
    """
    profile_yml = PROJECT_ROOT / "profile.yml"
    if not profile_yml.exists():
        return "wellness"
    try:
        import yaml
        data = yaml.safe_load(profile_yml.read_text(encoding="utf-8")) or {}
        return str(data.get("coach_profile") or "wellness").strip()
    except Exception:
        return "wellness"


def _load_blood_reference_ranges(path: Path = BLOOD_RANGES_YML) -> dict:
    """Load the blood reference-ranges YAML. Empty dict if missing —
    interpretation degrades to 'sin_rango' for every marker."""
    if not Path(path).exists():
        return {}
    try:
        import yaml
        data = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"  ! could not read {path.name}: {e}")
        return {}


def _target_band(spec: dict, profile: str) -> tuple[float, float] | None:
    """Resolve the soft-flag band for a marker spec + active profile.

    Precedence: spec['target_atleta'][profile] → ['default'] → None.
    """
    t = spec.get("target_atleta")
    if not isinstance(t, dict):
        return None
    band = t.get(profile) or t.get("default")
    if isinstance(band, list) and len(band) == 2:
        return float(band[0]), float(band[1])
    return None


def _classify_value(value: float, spec: dict, profile: str) -> tuple[str, str, str | None]:
    """Given a value + marker spec + active profile, return:

        (estado, flag_severity, trigger_msg)

    where:
        estado          ∈ {'bajo','borderline_bajo','normal','borderline_alto',
                           'alto','sin_rango'}
        flag_severity   ∈ {'dura','blanda','none'}
        trigger_msg     = the matching string from spec['trigger_si'] if any,
                          else None.

    Hard flags (`dura`) come from `trigger_si` and are universal.
    Soft flags (`blanda`) fire when the value falls outside the
    profile-specific target band but inside the lab range.
    """
    rango = spec.get("rango_lab")
    if not (isinstance(rango, list) and len(rango) == 2):
        return "sin_rango", "none", None
    # `null` (None) marks an "open" side of the range — no flag/borderline
    # in that direction. Used e.g. for HDL ("≥ 40, no upper bound") and
    # cholesterol total ("no lower bound").
    lo = float(rango[0]) if rango[0] is not None else None
    hi = float(rango[1]) if rango[1] is not None else None

    # --- hard flags first (trigger_si evaluation) ---
    trigger_msg = None
    for rule in spec.get("trigger_si") or []:
        rule_s = str(rule)
        # crude but transparent: each rule string contains its own threshold,
        # we evaluate just the inequalities we know how to parse. Pattern:
        # "valor [<|<=|>|>=] N ..."
        m = re.search(r"valor\s*(<=|>=|<|>)\s*([\d.]+)", rule_s)
        if not m:
            continue
        op, thr_s = m.group(1), m.group(2)
        try:
            thr = float(thr_s)
        except ValueError:
            continue
        fires = (
            (op == "<"  and value <  thr) or
            (op == "<=" and value <= thr) or
            (op == ">"  and value >  thr) or
            (op == ">=" and value >= thr)
        )
        if fires:
            trigger_msg = rule_s
            break

    # --- estado vs lab range ---
    # When a side is `None`, that direction is open (no out-of-range and
    # no borderline classification on that side).
    if lo is not None and value < lo:
        estado = "bajo"
    elif hi is not None and value > hi:
        estado = "alto"
    else:
        if lo is not None and hi is not None:
            span = max(hi - lo, 1e-9)
            border = 0.10 * span
            if value <= lo + border:
                estado = "borderline_bajo"
            elif value >= hi - border:
                estado = "borderline_alto"
            else:
                estado = "normal"
        elif lo is not None:
            # open above — only check borderline_bajo (10% of lo)
            estado = "borderline_bajo" if value <= lo * 1.10 else "normal"
        elif hi is not None:
            # open below — only check borderline_alto (10% of hi)
            estado = "borderline_alto" if value >= hi * 0.90 else "normal"
        else:
            estado = "normal"

    # --- soft flag from target_atleta band ---
    if trigger_msg is not None:
        severity = "dura"
    else:
        band = _target_band(spec, profile)
        if band is not None:
            in_band = True
            if band[0] is not None and value < band[0]:
                in_band = False
            if band[1] is not None and value > band[1]:
                in_band = False
            severity = "none" if in_band else "blanda"
        elif estado in ("bajo", "alto"):
            # out of lab range but no trigger configured — treat as soft
            severity = "blanda"
        else:
            severity = "none"

    return estado, severity, trigger_msg


def _compute_trend(values: list[float]) -> str:
    """Heuristic trend tag from chronologically-sorted values.

    Returns one of:
        'sin_historia' — fewer than 2 points
        'estable'      — last 4 vary < 10% of mean and no sustained direction
        'subiendo'     — last value > penúltimo AND > avg of previous two
        'bajando'      — symmetric
        'volatil'      — last 4 have ≥ 2 sign changes AND range > 20% of mean
    """
    vs = [v for v in values if v is not None]
    if len(vs) < 2:
        return "sin_historia"
    tail = vs[-4:]
    mean = sum(tail) / len(tail) or 1e-9
    rng = max(tail) - min(tail)

    # volátil
    if len(tail) >= 3:
        deltas = [tail[i+1] - tail[i] for i in range(len(tail) - 1)]
        sign_changes = sum(
            1 for i in range(len(deltas) - 1)
            if (deltas[i] > 0) != (deltas[i+1] > 0) and deltas[i] != 0 and deltas[i+1] != 0
        )
        if sign_changes >= 2 and rng > 0.20 * abs(mean):
            return "volatil"

    last = tail[-1]
    prev = tail[-2]
    prev_avg = sum(tail[:-2]) / max(len(tail) - 2, 1) if len(tail) >= 3 else prev

    if rng < 0.10 * abs(mean):
        return "estable"
    if last > prev and last > prev_avg:
        return "subiendo"
    if last < prev and last < prev_avg:
        return "bajando"
    return "estable"


def interpret_blood_panel(
    path: Path = BLOOD_XLSX,
    ranges_path: Path = BLOOD_RANGES_YML,
    profile: str | None = None,
) -> dict:
    """Read blood.xlsx + ranges YAML, compute the interpretation of the
    most-recent draw + per-marker trends.

    Returns:
        {
          'fecha_ultima':    date,
          'profile':         str,
          'fechas_todas':    list[str],
          'flags_duras':     list[interpreted_row],
          'flags_blandas':   list[interpreted_row],
          'estables':        list[interpreted_row],
          'sin_interpretacion': list[raw_row + tendencia],
          'todos':           list[interpreted_row] (last draw, every marker),
          'por_categoria':   dict[categoria, list[interpreted_row]],
          'historico':       list[dict] (every draw, every marker — for the
                             "Histórico completo" section of the .md),
        }

    Each interpreted_row carries:
        marker, valor, unidad, rango_lab, target_band, estado,
        flag_severity, trigger_msg, tendencia, lectura, categoria.
    """
    if profile is None:
        profile = _active_coach_profile()

    all_rows = load_blood_panel(path)
    if not all_rows:
        return {
            "fecha_ultima": None, "profile": profile, "fechas_todas": [],
            "flags_duras": [], "flags_blandas": [], "estables": [],
            "sin_interpretacion": [], "todos": [], "por_categoria": {},
            "historico": [],
        }
    ranges = _load_blood_reference_ranges(ranges_path)
    fechas = sorted({r["fecha"] for r in all_rows})
    latest_iso = fechas[-1]
    latest_rows = [r for r in all_rows if r["fecha"] == latest_iso]

    # Build per-marker history (ascending) for trend computation
    history_by_marker: dict[str, list[float]] = {}
    for r in all_rows:
        history_by_marker.setdefault(r["marker"], []).append((r["fecha"], r["valor"]))
    for m in history_by_marker:
        history_by_marker[m] = [v for _, v in sorted(history_by_marker[m])]

    interpreted: list[dict] = []
    for r in latest_rows:
        marker = r["marker"]
        spec = ranges.get(marker) or {}
        history = history_by_marker.get(marker, [])
        tendencia = _compute_trend(history)

        if spec:
            estado, severity, trigger_msg = _classify_value(r["valor"], spec, profile)
            lectura_block = spec.get("lectura_entrenamiento") or {}
            if estado in ("bajo", "borderline_bajo"):
                lectura = lectura_block.get("bajo") or lectura_block.get("en_rango") or ""
            elif estado in ("alto", "borderline_alto"):
                lectura = lectura_block.get("alto") or lectura_block.get("en_rango") or ""
            else:
                lectura = lectura_block.get("en_rango") or ""
            interpreted.append({
                "marker": marker,
                "valor": r["valor"],
                "unidad": spec.get("unidad", ""),
                "rango_lab": spec.get("rango_lab"),
                "target_band": _target_band(spec, profile),
                "estado": estado,
                "flag_severity": severity,
                "trigger_msg": trigger_msg,
                "tendencia": tendencia,
                "lectura": lectura,
                "categoria": spec.get("categoria") or "otros",
            })
        else:
            interpreted.append({
                "marker": marker,
                "valor": r["valor"],
                "unidad": "",
                "rango_lab": None,
                "target_band": None,
                "estado": "sin_rango",
                "flag_severity": "none",
                "trigger_msg": None,
                "tendencia": tendencia,
                "lectura": "",
                "categoria": "sin_interpretacion",
            })

    flags_duras   = [x for x in interpreted if x["flag_severity"] == "dura"]
    flags_blandas = [x for x in interpreted if x["flag_severity"] == "blanda"]
    estables      = [x for x in interpreted if x["flag_severity"] == "none"
                                              and x["categoria"] != "sin_interpretacion"]
    sin_interp    = [x for x in interpreted if x["categoria"] == "sin_interpretacion"]

    por_categoria: dict[str, list[dict]] = {}
    for x in interpreted:
        if x["categoria"] == "sin_interpretacion":
            continue
        por_categoria.setdefault(x["categoria"], []).append(x)
    for cat in por_categoria:
        por_categoria[cat].sort(key=lambda x: x["marker"].lower())

    return {
        "fecha_ultima": date.fromisoformat(latest_iso),
        "profile": profile,
        "fechas_todas": fechas,
        "flags_duras": flags_duras,
        "flags_blandas": flags_blandas,
        "estables": estables,
        "sin_interpretacion": sin_interp,
        "todos": interpreted,
        "por_categoria": por_categoria,
        "historico": all_rows,
    }


# --- anthropometry: heurísticas inline ---------------------------------

# Targets de body composition por perfil. `None` = no flag automático.
_ANTHRO_TARGETS: dict[str, dict[str, tuple[float, float] | None]] = {
    # (low, high) inclusive
    "Masa Adiposa (%)": {
        "hyrox": (10.0, 15.0),
        "default": (10.0, 18.0),
        "wellness": (10.0, 22.0),
    },
    "FFMI": {
        # FFMI < 18 = subdesarrollado, > 25 = sospecha PEDs. Para Hyrox 20-24 es la banda sólida.
        "hyrox": (20.0, 24.0),
        "default": (18.0, 25.0),
    },
}

# Etiquetas humanas por prefijo de nombre — el Excel mezcla varios protocolos.
_ANTHRO_PREFIX_BLOCK = [
    ("Pl. ",     "Pliegues"),
    ("Pr. ",     "Perímetros"),
    ("Diam. ",   "Diámetros"),
    ("Masa ",    "Masas corporales"),
    ("% Adiposidad ", "Distribución adiposa"),
    ("Sum. ",    "Sumatorias"),
    ("Endomorfia",  "Somatotipo"),
    ("Mesomorfia",  "Somatotipo"),
    ("Ectomorfia",  "Somatotipo"),
    ("S.D.D",    "Somatotipo"),
    ("Area Muscular ", "Áreas musculares"),
    ("Req. ",    "Requerimiento energético"),
    ("FFMI",     "Índices de desarrollo"),
    ("Ind. ",    "Índices de desarrollo"),
    ("Peso ",    "Morfología global"),
    ("Talla",    "Morfología global"),
    ("Edad ",    "Morfología global"),
]


def _anthro_block(variable: str) -> str:
    for prefix, label in _ANTHRO_PREFIX_BLOCK:
        if variable.startswith(prefix) or variable == prefix.strip():
            return label
    return "Otros"


def interpret_anthropometry(
    path: Path = ANTHRO_XLSX,
    profile: str | None = None,
) -> dict:
    """Read anthropometry.xlsx + apply inline heuristics, returning the
    interpretation of the most-recent eval + trends.

    Special cases:
        - `Masa Muscular (kg)` con valor > 200 → flag explícito 'unit_bug'
          (preserva el valor crudo y excluye ese punto del cómputo de
          tendencia).

    Returns same shape as `interpret_blood_panel` but keyed on
    `variable` instead of `marker`, and adds:
        - `targets_profile`: dict de targets que se usaron (para que la
          .md pueda explicitar la banda usada).
    """
    if profile is None:
        profile = _active_coach_profile()

    all_rows = load_anthropometry(path)
    if not all_rows:
        return {
            "fecha_ultima": None, "profile": profile, "fechas_todas": [],
            "flags_duras": [], "flags_blandas": [], "estables": [],
            "todos": [], "por_bloque": {}, "historico": [],
            "targets_profile": {},
        }
    fechas = sorted({r["fecha"] for r in all_rows})
    latest_iso = fechas[-1]
    latest_rows = [r for r in all_rows if r["fecha"] == latest_iso]

    # history per variable, excluding unit-bug points for Masa Muscular (kg)
    history_by_var: dict[str, list[tuple[str, float]]] = {}
    for r in all_rows:
        v = r["valor"]
        if r["variable"] == "Masa Muscular (kg)" and v is not None and v > 200:
            continue
        history_by_var.setdefault(r["variable"], []).append((r["fecha"], v))
    for k in history_by_var:
        history_by_var[k] = [v for _, v in sorted(history_by_var[k])]

    targets_used: dict[str, tuple[float, float]] = {}
    interpreted: list[dict] = []
    for r in latest_rows:
        var = r["variable"]
        val = r["valor"]
        tendencia = _compute_trend(history_by_var.get(var, []))
        bloque = _anthro_block(var)

        flag_severity = "none"
        trigger_msg = None
        lectura = ""

        # Unit-bug check
        if var == "Masa Muscular (kg)" and val is not None and val > 200:
            flag_severity = "dura"
            trigger_msg = f"unit_bug: valor crudo {val} > 200 kg — probable error de unidad en la fuente; no se computa tendencia con este punto"
            lectura = "Dato crudo preservado pero excluido de la tendencia hasta que se corrija en el Excel."

        # Target-band check (soft flag)
        band_spec = _ANTHRO_TARGETS.get(var)
        band = None
        if band_spec:
            band = band_spec.get(profile) or band_spec.get("default")
            if band:
                targets_used[var] = band
                if flag_severity == "none" and not (band[0] <= val <= band[1]):
                    flag_severity = "blanda"

        interpreted.append({
            "variable": var,
            "valor": val,
            "estado": ("alto"  if band and val > band[1]
                       else "bajo" if band and val < band[0]
                       else "normal" if band
                       else "sin_rango"),
            "flag_severity": flag_severity,
            "trigger_msg": trigger_msg,
            "tendencia": tendencia,
            "lectura": lectura,
            "bloque": bloque,
            "target_band": band,
        })

    flags_duras   = [x for x in interpreted if x["flag_severity"] == "dura"]
    flags_blandas = [x for x in interpreted if x["flag_severity"] == "blanda"]
    estables      = [x for x in interpreted if x["flag_severity"] == "none"]

    por_bloque: dict[str, list[dict]] = {}
    for x in interpreted:
        por_bloque.setdefault(x["bloque"], []).append(x)
    for k in por_bloque:
        por_bloque[k].sort(key=lambda x: x["variable"].lower())

    return {
        "fecha_ultima": date.fromisoformat(latest_iso),
        "profile": profile,
        "fechas_todas": fechas,
        "flags_duras": flags_duras,
        "flags_blandas": flags_blandas,
        "estables": estables,
        "todos": interpreted,
        "por_bloque": por_bloque,
        "historico": all_rows,
        "targets_profile": targets_used,
    }


# --- .md renderers -------------------------------------------------------

_TREND_ARROW = {
    "subiendo":    "↑",
    "bajando":     "↓",
    "estable":     "→",
    "volatil":     "↕",
    "sin_historia": "·",
}

_ESTADO_LABEL = {
    "bajo":             "bajo",
    "borderline_bajo":  "borderline-bajo",
    "normal":           "normal",
    "borderline_alto":  "borderline-alto",
    "alto":             "alto",
    "sin_rango":        "—",
}


def _fmt_value(v: float) -> str:
    if v is None:
        return "—"
    if isinstance(v, float) and not v.is_integer() and abs(v) < 1000:
        return f"{v:g}"
    return f"{v}"


def _fmt_band(band) -> str:
    if not band:
        return "—"
    lo, hi = band
    if lo is None and hi is None:
        return "—"
    if lo is None:
        return f"≤ {_fmt_value(hi)}"
    if hi is None:
        return f"≥ {_fmt_value(lo)}"
    return f"{_fmt_value(lo)}–{_fmt_value(hi)}"


def _render_blood_panel_md(interp: dict) -> str:
    """Render the full blood_panel.md from the dict returned by
    `interpret_blood_panel()`."""
    if interp["fecha_ultima"] is None:
        return (
            "# Panel sanguíneo — historial e interpretación\n\n"
            "_No hay datos en `data/manual/blood.xlsx`._\n"
        )

    profile = interp["profile"]
    last = interp["fecha_ultima"].isoformat()
    fduras   = interp["flags_duras"]
    fblandas = interp["flags_blandas"]
    estables = interp["estables"]
    sin_int  = interp["sin_interpretacion"]
    por_cat  = interp["por_categoria"]
    historico = interp["historico"]
    fechas_todas = interp["fechas_todas"]

    out = [
        "# Panel sanguíneo — historial e interpretación",
        "",
        "> Generado por `_session_lib.refresh_blood_panel_md()` a partir de",
        "> `data/manual/blood.xlsx` + `data/manual/blood_reference_ranges.yml`.",
        "> **No editar a mano.** Las extracciones se gestionan agregando filas",
        "> al Excel; este archivo se regenera automáticamente cuando el Excel",
        "> está más fresco (ver CLAUDE.md §0.4-bis).",
        ">",
        f"> Última regeneración: {datetime.utcnow().isoformat(timespec='seconds')}Z",
        f"> Perfil activo: **{profile}** · extracciones totales: {len(fechas_todas)}",
        "",
        "---",
        "",
        f"## Estado actual (al {last} — última extracción)",
        "",
        f"**Resumen ejecutivo:** {len(interp['todos'])} markers medidos · "
        f"{len(fduras)} flag(s) dura(s) · {len(fblandas)} flag(s) blanda(s).",
        "",
        "### 🔴 Flags duras",
        "",
    ]
    if not fduras:
        out.append("_Sin flags duras en esta extracción._")
    else:
        for x in sorted(fduras, key=lambda y: y["marker"].lower()):
            out.append(
                f"- **{x['marker']} = {_fmt_value(x['valor'])} {x['unidad']}** "
                f"(rango lab {_fmt_band(x['rango_lab'])}"
                + (f", target atleta {_fmt_band(x['target_band'])}" if x['target_band'] else "")
                + f") → **{_ESTADO_LABEL[x['estado']]}**. "
                f"Tendencia: {_TREND_ARROW[x['tendencia']]} {x['tendencia']}."
            )
            if x.get("trigger_msg"):
                out.append(f"  - {x['trigger_msg']}")
            if x.get("lectura"):
                out.append(f"  - *Lectura entrenamiento:* {x['lectura']}")
    out.append("")
    out.append("### 🟡 Flags blandas / contextuales")
    out.append("")
    if not fblandas:
        out.append("_Sin flags blandas en esta extracción._")
    else:
        for x in sorted(fblandas, key=lambda y: y["marker"].lower()):
            band = (f", target atleta {_fmt_band(x['target_band'])}" if x['target_band'] else "")
            out.append(
                f"- **{x['marker']} = {_fmt_value(x['valor'])} {x['unidad']}** "
                f"(rango lab {_fmt_band(x['rango_lab'])}"
                + band
                + f") → {_ESTADO_LABEL[x['estado']]}. "
                f"Tendencia: {_TREND_ARROW[x['tendencia']]} {x['tendencia']}."
            )
            if x.get("lectura"):
                out.append(f"  - *Lectura:* {x['lectura']}")
    out.append("")
    out.append("### ✅ Estables / sin flag")
    out.append("")
    if not estables:
        out.append("_Sin markers estables interpretados (¿la YAML está vacía?)._")
    else:
        chunks = [
            f"{x['marker']} = {_fmt_value(x['valor'])}"
            + (f" {x['unidad']}" if x['unidad'] else "")
            for x in sorted(estables, key=lambda y: y["marker"].lower())
        ]
        # group into shortish lines
        out.append(", ".join(chunks))
    out.append("")
    out.append("---")
    out.append("")
    out.append(f"## Por categoría — última extracción ({last})")
    out.append("")
    cat_order = ["hemograma", "ferroso", "metabolico", "lipidico",
                 "hepatico", "renal", "endocrino", "vitaminas", "iones", "otros"]
    cat_label = {
        "hemograma": "Hemograma",
        "ferroso":   "Perfil ferroso",
        "metabolico":"Metabolismo glucémico",
        "lipidico":  "Perfil lipídico",
        "hepatico":  "Función hepática",
        "renal":     "Función renal",
        "endocrino": "Endócrino",
        "vitaminas": "Vitaminas",
        "iones":     "Iones / electrolitos",
        "otros":     "Otros",
    }
    for cat in cat_order:
        rows = por_cat.get(cat) or []
        if not rows:
            continue
        out.append(f"### {cat_label[cat]}")
        out.append("")
        out.append("| Marker | Valor | Rango lab | Target atleta | Estado | Tendencia |")
        out.append("|---|---:|---|---|---|---|")
        for x in rows:
            out.append(
                f"| {x['marker']} | {_fmt_value(x['valor'])} {x['unidad']} "
                f"| {_fmt_band(x['rango_lab'])} "
                f"| {_fmt_band(x['target_band'])} "
                f"| {_ESTADO_LABEL[x['estado']]} "
                f"| {_TREND_ARROW[x['tendencia']]} {x['tendencia']} |"
            )
        out.append("")

    if sin_int:
        out.append("### Otros markers medidos (sin interpretación configurada)")
        out.append("")
        out.append("| Marker | Valor | Tendencia |")
        out.append("|---|---:|---|")
        for x in sorted(sin_int, key=lambda y: y["marker"].lower()):
            out.append(
                f"| {x['marker']} | {_fmt_value(x['valor'])} "
                f"| {_TREND_ARROW[x['tendencia']]} {x['tendencia']} |"
            )
        out.append("")
        out.append(
            "_Estos markers aparecen en el Excel pero no tienen entrada en "
            "`blood_reference_ranges.yml`. Agregalos al YAML para activar "
            "interpretación + banderas._"
        )
        out.append("")

    out.append("---")
    out.append("")
    out.append("## Histórico completo — append-only")
    out.append("")
    # group historico by fecha (most recent first)
    by_fecha: dict[str, list[dict]] = {}
    for r in historico:
        by_fecha.setdefault(r["fecha"], []).append(r)
    for fecha in sorted(by_fecha.keys(), reverse=True):
        out.append(f"### Extracción {fecha}")
        out.append("")
        out.append("| Marker | Valor | Sheet |")
        out.append("|---|---:|---|")
        for r in sorted(by_fecha[fecha], key=lambda y: y["marker"].lower()):
            out.append(
                f"| {r['marker']} | {_fmt_value(r['valor'])} | {r.get('sheet','')} |"
            )
        out.append("")
    return "\n".join(out).rstrip() + "\n"


def _render_body_composition_md(interp: dict) -> str:
    """Render the full body_composition.md from `interpret_anthropometry()`."""
    if interp["fecha_ultima"] is None:
        return (
            "# Composición corporal — historial e interpretación\n\n"
            "_No hay datos en `data/manual/anthropometry.xlsx`._\n"
        )

    profile = interp["profile"]
    last = interp["fecha_ultima"].isoformat()
    fduras   = interp["flags_duras"]
    fblandas = interp["flags_blandas"]
    estables = interp["estables"]
    por_bloque = interp["por_bloque"]
    historico = interp["historico"]
    fechas_todas = interp["fechas_todas"]
    targets = interp["targets_profile"]

    # locate key headline metrics in the latest eval
    by_var = {x["variable"]: x for x in interp["todos"]}
    headline_keys = [
        "Peso (kg)",
        "Masa Adiposa (%)",
        "Masa Adiposa (kg)",
        "Masa Muscular (%)",
        "Masa Muscular (kg)",
        "FFMI",
        "FFMI Normalizado",
        "Sum. de 6 Plieg. (mm)",
        "Pr. Umbilical (cm)",
    ]

    out = [
        "# Composición corporal — historial e interpretación",
        "",
        "> Generado por `_session_lib.refresh_body_composition_md()` a partir",
        "> de `data/manual/anthropometry.xlsx`. **No editar a mano.** Las",
        "> evaluaciones se gestionan agregando columnas al Excel; este",
        "> archivo se regenera automáticamente.",
        ">",
        f"> Última regeneración: {datetime.utcnow().isoformat(timespec='seconds')}Z",
        f"> Perfil activo: **{profile}** · evals totales: {len(fechas_todas)}",
        "",
        "---",
        "",
        f"## Estado actual (al {last} — última eval)",
        "",
        "**Resumen ejecutivo:**",
        "",
    ]
    for key in headline_keys:
        x = by_var.get(key)
        if not x:
            continue
        band_str = ""
        if x["target_band"]:
            band_str = f" (banda perfil {profile}: {_fmt_band(x['target_band'])})"
        out.append(
            f"- **{key}** = {_fmt_value(x['valor'])} · "
            f"{_TREND_ARROW[x['tendencia']]} {x['tendencia']}{band_str}"
        )
    out.append("")

    out.append("### 🔴 Flags duras")
    out.append("")
    if not fduras:
        out.append("_Sin flags duras en esta eval._")
    else:
        for x in fduras:
            out.append(f"- **{x['variable']} = {_fmt_value(x['valor'])}** → {x['trigger_msg']}")
            if x["lectura"]:
                out.append(f"  - {x['lectura']}")
    out.append("")
    out.append("### 🟡 Flags blandas")
    out.append("")
    if not fblandas:
        out.append("_Sin flags blandas en esta eval._")
    else:
        for x in fblandas:
            band = f" (target {profile}: {_fmt_band(x['target_band'])})" if x["target_band"] else ""
            out.append(
                f"- **{x['variable']} = {_fmt_value(x['valor'])}**{band} → "
                f"{_ESTADO_LABEL[x['estado']]}. "
                f"Tendencia: {_TREND_ARROW[x['tendencia']]} {x['tendencia']}."
            )
    out.append("")

    # trayectoria 12m de los headline
    out.append("### Trayectoria headline (todas las evals)")
    out.append("")
    out.append("| Variable | Trayectoria (más antigua → más reciente) |")
    out.append("|---|---|")
    for key in headline_keys:
        hist = anthropometry_variable_history(key)
        if not hist:
            continue
        cells = ", ".join(f"{d}: {_fmt_value(v)}" for d, v in hist)
        out.append(f"| {key} | {cells} |")
    out.append("")

    out.append("---")
    out.append("")
    out.append(f"## Por bloque — última eval ({last})")
    out.append("")
    bloque_order = [
        "Morfología global", "Índices de desarrollo", "Pliegues", "Sumatorias",
        "Perímetros", "Diámetros", "Masas corporales", "Distribución adiposa",
        "Somatotipo", "Áreas musculares", "Requerimiento energético", "Otros",
    ]
    for bloque in bloque_order:
        rows = por_bloque.get(bloque) or []
        if not rows:
            continue
        out.append(f"### {bloque}")
        out.append("")
        out.append("| Variable | Valor | Estado | Target perfil | Tendencia |")
        out.append("|---|---:|---|---|---|")
        for x in rows:
            out.append(
                f"| {x['variable']} | {_fmt_value(x['valor'])} "
                f"| {_ESTADO_LABEL[x['estado']]} "
                f"| {_fmt_band(x['target_band'])} "
                f"| {_TREND_ARROW[x['tendencia']]} {x['tendencia']} |"
            )
        out.append("")

    out.append("---")
    out.append("")
    out.append("## Histórico completo — append-only")
    out.append("")
    by_fecha: dict[str, list[dict]] = {}
    for r in historico:
        by_fecha.setdefault(r["fecha"], []).append(r)
    for fecha in sorted(by_fecha.keys(), reverse=True):
        out.append(f"### Eval {fecha}")
        out.append("")
        out.append("| Variable | Valor |")
        out.append("|---|---:|")
        for r in sorted(by_fecha[fecha], key=lambda y: y["variable"].lower()):
            out.append(f"| {r['variable']} | {_fmt_value(r['valor'])} |")
        out.append("")
    return "\n".join(out).rstrip() + "\n"


def refresh_blood_panel_md(
    output: Path = BLOOD_PANEL_MD,
    profile: str | None = None,
) -> Path:
    """Regenerate `blood_panel.md` from `data/manual/blood.xlsx` +
    `data/manual/blood_reference_ranges.yml`. Idempotent.

    Returns the output path. Profile defaults to the active coach_profile.
    """
    interp = interpret_blood_panel(profile=profile)
    out = Path(output)
    out.write_text(_render_blood_panel_md(interp), encoding="utf-8")
    return out


def refresh_body_composition_md(
    output: Path = BODY_COMP_MD,
    profile: str | None = None,
) -> Path:
    """Regenerate `body_composition.md` from
    `data/manual/anthropometry.xlsx`. Idempotent."""
    interp = interpret_anthropometry(profile=profile)
    out = Path(output)
    out.write_text(_render_body_composition_md(interp), encoding="utf-8")
    return out


def manual_data_is_stale() -> dict[str, bool]:
    """Compare mtimes: is the Excel newer than the rendered .md?

    Returns:
        {'blood': True|False, 'anthropometry': True|False}
        True means the .md needs regeneration.
    """
    def stale(src: Path, md: Path) -> bool:
        if not src.exists():
            return False
        if not md.exists():
            return True
        return src.stat().st_mtime > md.stat().st_mtime
    return {
        "blood":         stale(BLOOD_XLSX, BLOOD_PANEL_MD),
        "anthropometry": stale(ANTHRO_XLSX, BODY_COMP_MD),
    }
